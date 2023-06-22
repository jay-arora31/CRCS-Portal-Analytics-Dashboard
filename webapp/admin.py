from django.contrib import admin
from django.http import HttpResponseRedirect
from django.shortcuts import render
from openpyxl import load_workbook
from .models import CooperativeSociety

class CooperativeSocietyAdmin(admin.ModelAdmin):
    actions = ['import_data']
    
    def import_data(self, request, queryset):
        if 'apply' in request.POST:
            # Get the uploaded file from the request
            file = request.FILES.get('file')
            
            # Load the Excel workbook
            wb = load_workbook(file)
            sheet = wb.active
            
            # Iterate over rows and create objects
            for row in sheet.iter_rows(values_only=True):
                name = row[0]
                address = row[1]
                state = row[2]
                district = row[3]
                registration_date = row[4]
                area_of_operation = row[5]
                sector_type = row[6]

                # Create a new CooperativeSociety object
                society = CooperativeSociety(
                    name=name,
                    address=address,
                    state=state,
                    district=district,
                    registration_date=registration_date,
                    area_of_operation=area_of_operation,
                    sector_type=sector_type
                )

                # Save the object
                society.save()
            
            self.message_user(request, 'Data imported successfully.')
            return HttpResponseRedirect(request.get_full_path())
        
        context = {
            'title': 'Import Data',
            'opts': self.model._meta,
            'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
            'media': self.media,
            'queryset': queryset,
            'app_label': self.model._meta.app_label,
        }
        
        return render(request, 'admin/import_data.html', context)
    
    import_data.short_description = 'Import data from Excel'

admin.site.register(CooperativeSociety, CooperativeSocietyAdmin)
